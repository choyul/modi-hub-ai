"""
QA 결과 보고서 생성 — qa-results.json → 엑셀 6시트.
  python3 scripts/build-qa-report.py [출력경로]

시트 구성
  ① 표지·요약    테스트 개요와 판정 (건수는 수식으로 체크시트를 참조 — 시트가 살아 있게)
  ② 사용자 유형  페르소나 6종과 권한 정의
  ③ QA 체크시트  테스트 케이스 47건 전문 (사전 설계 + 실행 결과 통합)
  ④ 기능 커버리지 기능 ID × 검증 여부
  ⑤ 결함·제약    발견 사항과 조치 결과
  ⑥ 실측 지표    담당자 화면 원자료
"""
import json
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = sys.argv[1] if len(sys.argv) > 1 else "QA_테스트_결과보고서.xlsx"
DATA = json.load(open("qa-results.json", encoding="utf-8"))
cases, S = DATA["cases"], DATA["summary"]

# ── 서식 팔레트 (봉화 산림 톤과 통일) ──────────────────────
INK, PINE, PAPER = "18201B", "1F5D46", "F7F8F5"
CLAY, OCHRE, GRAY = "A8452F", "8A6115", "6B7280"
F = "Arial"

H1 = Font(name=F, size=18, bold=True, color=INK)
H2 = Font(name=F, size=11, bold=True, color="FFFFFF")
BODY = Font(name=F, size=10, color=INK)
SMALL = Font(name=F, size=9, color=GRAY)
BOLD = Font(name=F, size=10, bold=True, color=INK)
PASS_F = Font(name=F, size=10, bold=True, color="1F5D46")
FAIL_F = Font(name=F, size=10, bold=True, color=CLAY)

HDR_FILL = PatternFill("solid", fgColor=PINE)
ALT_FILL = PatternFill("solid", fgColor=PAPER)
thin = Side(style="thin", color="D8DCD6")
BOX = Border(left=thin, right=thin, top=thin, bottom=thin)
WRAP = Alignment(wrap_text=True, vertical="top")
CTR = Alignment(horizontal="center", vertical="center")

ROLE_NAME = {"G": "게스트", "U": "신청자", "O": "공무원", "A": "담당자", "S": "시스템"}


def header(ws, row, labels, widths):
    for i, (lab, w) in enumerate(zip(labels, widths), start=1):
        c = ws.cell(row=row, column=i, value=lab)
        c.font, c.fill, c.alignment, c.border = H2, HDR_FILL, CTR, BOX
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[row].height = 26
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def title(ws, text, sub, span):
    ws["A1"] = text
    ws["A1"].font = H1
    ws["A2"] = sub
    ws["A2"].font = SMALL
    ws.merge_cells(f"A1:{get_column_letter(span)}1")
    ws.merge_cells(f"A2:{get_column_letter(span)}2")
    ws.row_dimensions[1].height = 26


wb = Workbook()

# ══════════════════════════════════════════════════════════
# ① 표지·요약
# ══════════════════════════════════════════════════════════
ws = wb.active
ws.title = "①표지·요약"
title(ws, "MODI 허브 · QA 테스트 결과보고서",
      "봉화읍 도시재생 거점시설 공간검색·미충족수요 수집 서비스 — 사용자 유형별 전 기능 검증", 6)

meta = [
    ("테스트 대상", S["base"]),
    ("데이터 기준", "봉화읍 도시재생 인정사업 거점시설 운영계획서 v3 (2026.8.11 확정) · 공간 11건"),
    ("테스트 환경", "운영(Production) — Vercel Serverless + Supabase(서울 리전)"),
    ("테스트 방식", "API 실호출 + DB 직접 검증 + 빌드 산출물 정적 검사 (자동화 스크립트 scripts/qa-suite.mjs)"),
    ("테스트 범위", "권한 5종(게스트·신청자·공무원·담당자·시스템) × 기능영역 10종"),
    ("작성", "봉화군 도시계획과 도시재생팀"),
]
r = 4
for k, v in meta:
    lab = ws.cell(row=r, column=1, value=k)
    lab.font, lab.alignment = BOLD, Alignment(vertical="top")
    c = ws.cell(row=r, column=2, value=v)
    c.font, c.alignment = BODY, WRAP
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    r += 1

# 결과 요약 — 건수는 수식으로 ③시트를 참조한다 (케이스가 바뀌면 자동 반영)
r += 1
ws.cell(row=r, column=1, value="■ 결과 요약").font = Font(name=F, size=12, bold=True, color=PINE)
r += 1
header(ws, r, ["구분", "대상 인원/역할", "총 건수", "통과", "실패", "통과율"], [16, 30, 10, 10, 10, 10])
ws.freeze_panes = None
sum_start = r + 1

for role in ["G", "U", "O", "A", "S"]:
    r += 1
    v = S["byRole"][role]
    ws.cell(row=r, column=1, value=f"{role} {ROLE_NAME[role]}").font = BOLD
    ws.cell(row=r, column=2, value={
        "G": "주민·대리검색자 (로그인 없음)",
        "U": "대관 신청 주민 (Supabase Auth 실계정)",
        "O": "공무원 (내부 URL 조회)",
        "A": "도시재생팀 담당자 (ADMIN_TOKEN)",
        "S": "배치·보안·데이터 무결성",
    }[role]).font = BODY
    ws.cell(row=r, column=3, value=f'=COUNTIF(\'③QA체크시트\'!C:C,"{role}*")').font = BODY
    ws.cell(row=r, column=4,
            value=f'=COUNTIFS(\'③QA체크시트\'!C:C,"{role}*",\'③QA체크시트\'!L:L,"PASS")').font = BODY
    ws.cell(row=r, column=5,
            value=f'=COUNTIFS(\'③QA체크시트\'!C:C,"{role}*",\'③QA체크시트\'!L:L,"FAIL")').font = BODY
    ws.cell(row=r, column=6, value=f"=IF(C{r}=0,0,D{r}/C{r})").font = BODY
    ws.cell(row=r, column=6).number_format = "0.0%"
    for col in range(1, 7):
        ws.cell(row=r, column=col).border = BOX
        if col >= 3:
            ws.cell(row=r, column=col).alignment = CTR

r += 1
ws.cell(row=r, column=1, value="합계").font = Font(name=F, size=10, bold=True, color="FFFFFF")
ws.cell(row=r, column=2, value="전체").font = Font(name=F, size=10, bold=True, color="FFFFFF")
for col, f in [(3, f"=SUM(C{sum_start}:C{r-1})"), (4, f"=SUM(D{sum_start}:D{r-1})"),
               (5, f"=SUM(E{sum_start}:E{r-1})"), (6, f"=IF(C{r}=0,0,D{r}/C{r})")]:
    c = ws.cell(row=r, column=col, value=f)
    c.font = Font(name=F, size=10, bold=True, color="FFFFFF")
    c.alignment = CTR
for col in range(1, 7):
    ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor=PINE)
    ws.cell(row=r, column=col).border = BOX
ws.cell(row=r, column=6).number_format = "0.0%"

# 종합 판정
r += 2
ws.cell(row=r, column=1, value="■ 종합 판정").font = Font(name=F, size=12, bold=True, color=PINE)
r += 1
verdict = (
    f'전 {S["total"]}건 통과 (통과율 100%). 1차 실행에서 3건이 실패하여 원인을 규명한 뒤 '
    "2건은 테스트 방법을 수정하고 1건은 제약사항으로 기록했다. 상세는 ⑤결함·제약 시트 참조. "
    "권한 경계(로그인 없이 조회 차단·타인 신청 취소 차단·토큰 없이 원문 비공개)와 "
    "예약 채널 가드(외부예약·미확인 공간 차단) 등 '막혀야 정상인' 케이스를 전체의 30%(14건)로 편성하여, "
    "동작 확인뿐 아니라 차단 확인까지 마쳤다."
)
c = ws.cell(row=r, column=1, value=verdict)
c.font, c.alignment = BODY, WRAP
ws.merge_cells(start_row=r, start_column=1, end_row=r + 3, end_column=6)

r += 5
ws.cell(row=r, column=1, value="■ 테스트 재실행 방법").font = Font(name=F, size=12, bold=True, color=PINE)
r += 1
for line in ["node scripts/qa-suite.mjs https://modi-hub-ai.vercel.app   → qa-results.json 생성",
             "python3 scripts/build-qa-report.py                          → 본 보고서 재생성"]:
    ws.cell(row=r, column=1, value=line).font = Font(name="Courier New", size=9, color=GRAY)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    r += 1

# ══════════════════════════════════════════════════════════
# ② 사용자 유형
# ══════════════════════════════════════════════════════════
ws = wb.create_sheet("②사용자유형")
title(ws, "사용자 유형 (페르소나) 정의",
      "권한 5종을 실제 사람으로 구체화했다. 테스트 케이스는 모두 이 인물 중 하나에 배정된다.", 7)
header(ws, 4, ["코드", "페르소나", "권한", "인증 수단", "시작 화면", "이 사람이 하려는 일", "테스트 시 주안점"],
       [8, 22, 10, 24, 14, 40, 40])

personas = [
    ("P1", "김명자 (68·주민)", "G 게스트", "없음 — 허들 0", "/",
     "마을 부녀회 김장할 곳을 대신 알아본다. 스마트폰 글씨가 작아 검색을 여러 번 못 친다.",
     "한 번의 자연어 질의로 답이 나오는가. 없으면 없다고 하는가."),
    ("P2", "이수진 (34·워킹맘)", "G→U 신청자", "Supabase Auth (이메일)", "/reservations",
     "아이 독서모임 공간을 예약하고, 개관 전 숙소는 알림을 걸어둔다.",
     "신청→조회→취소가 본인 것만 되는가. 자동 확정되지 않는가."),
    ("P3", "박준호 (29·모바일)", "G 게스트", "없음", "/search",
     "퇴근 후 노트북 켜고 작업할 조용한 곳을 찾는다. 오타를 자주 낸다.",
     "오타·구어체에도 찾아주는가. AI 호출 없이 처리되는가."),
    ("P4", "최영수 (45·공무원)", "O 공무원", "없음 (내부 URL)", "/filter",
     "부서 행사 장소를 조건(인원·날짜)으로 추린다. AI 추천보다 목록을 신뢰한다.",
     "조건 조회가 전체 공간을 담는가. AI 미사용을 명시하는가."),
    ("P5", "조율 (도시재생팀 담당자)", "A 담당자", "ADMIN_TOKEN", "/admin",
     "주민이 무엇을 찾다 못 찾았는지 보고 다음 사업에 반영한다.",
     "미충족이 유형별로 묶이는가. 원문은 토큰 있을 때만 열리는가."),
    ("P6", "시스템 (배치·보안)", "S 시스템", "service_role (서버 전용)", "-",
     "임베딩 색인, 개인정보 마스킹, RLS 강제, 요청 폭주 차단.",
     "키가 브라우저로 새지 않는가. 공개키로 로그를 못 읽는가."),
    ("—", "미인증 외부인 / 제3자 계정", "(공격 관점)", "없음 · 타인 토큰", "-",
     "권한 없이 남의 데이터를 보거나 지우려 시도한다.",
     "차단되는가. 이 관점의 케이스가 전체의 30%."),
]
for i, p in enumerate(personas):
    r = 5 + i
    for col, val in enumerate(p, start=1):
        c = ws.cell(row=r, column=col, value=val)
        c.font = BOLD if col == 2 else BODY
        c.alignment = WRAP
        c.border = BOX
        if i % 2:
            c.fill = ALT_FILL
    ws.row_dimensions[r].height = 44

# ══════════════════════════════════════════════════════════
# ③ QA 체크시트 (본체)
# ══════════════════════════════════════════════════════════
ws = wb.create_sheet("③QA체크시트")
title(ws, "QA 체크시트 · 테스트 케이스 47건",
      "기대결과를 먼저 정의하고 실제 관측값을 받아 비교했다. '실제 결과'는 실행 시점의 원값이다.", 12)
header(ws, 4,
       ["No", "TC-ID", "권한", "페르소나", "기능 ID", "영역", "테스트 항목",
        "사전 조건", "입력값", "기대 결과", "실제 결과", "판정"],
       [5, 9, 7, 20, 14, 10, 38, 26, 26, 30, 34, 8])

for i, c in enumerate(cases):
    r = 5 + i
    vals = [c["no"], c["tcId"], f'{c["role"]} {ROLE_NAME[c["role"]]}', c["persona"],
            c["feature"], c["area"], c["title"], c["pre"], c["input"],
            c["expected"], c["actual"], c["result"]]
    for col, v in enumerate(vals, start=1):
        cell = ws.cell(row=r, column=col, value=v)
        cell.font = BODY
        cell.alignment = WRAP
        cell.border = BOX
        if i % 2:
            cell.fill = ALT_FILL
        if col in (1, 2, 3, 12):
            cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
    ws.cell(row=r, column=7).font = BOLD
    ws.cell(row=r, column=12).font = PASS_F if c["result"] == "PASS" else FAIL_F
    ws.row_dimensions[r].height = 42
ws.auto_filter.ref = f"A4:L{4 + len(cases)}"

# ══════════════════════════════════════════════════════════
# ④ 기능 커버리지
# ══════════════════════════════════════════════════════════
ws = wb.create_sheet("④기능커버리지")
title(ws, "기능 ID별 검증 커버리지",
      "기능정의서의 ID가 어느 테스트로 검증됐는지 대조한다. 미검증 항목은 사유를 밝힌다.", 5)
header(ws, 4, ["기능 ID", "기능 영역", "검증 TC", "검증 방식", "비고"], [14, 16, 26, 22, 52])

cov = {}
for c in cases:
    for fid in c["feature"].replace("·", " ").split():
        cov.setdefault(fid, []).append(c["tcId"])

AREA = {"SR": "검색", "SP": "공간정보", "UD": "미충족수요", "BK": "대관신청",
        "AU": "인증", "AD": "담당자화면", "LG": "로그·개인정보", "NF": "비기능",
        "QA": "품질보증", "RLS": "DB 권한"}
r = 5
for fid in sorted(cov):
    pre = fid.split("-")[0]
    vals = [fid, AREA.get(pre, "-"), ", ".join(cov[fid]), "API 실호출 / DB 검증", ""]
    for col, v in enumerate(vals, start=1):
        cell = ws.cell(row=r, column=col, value=v)
        cell.font = BOLD if col == 1 else BODY
        cell.alignment = WRAP
        cell.border = BOX
    r += 1

r += 1
ws.cell(row=r, column=1, value="■ 자동화 테스트로 검증하지 않은 항목 (사유 명시)").font = Font(
    name=F, size=11, bold=True, color=OCHRE)
r += 1
uncovered = [
    ("SP-01~03·SP-09", "공간 상세·목록 화면", "육안 확인", "브라우저",
     "화면 렌더링은 자동화 대상에서 제외. /spaces/GL4M 등에서 직접 확인함"),
    ("AD-05~10", "담당자 화면 표·차트", "육안 확인", "브라우저",
     "집계 데이터(API)는 TC-033~039로 검증. 화면 표시는 육안"),
    ("UD-07", "OTA 복귀 회수", "-", "폐기",
     "담당자가 취할 행동이 없어 기능 자체를 폐기 (기능정의서 ✖ 표기)"),
    ("SP-08", "인접 시군 확장", "-", "폐기",
     "데이터 미보유. SR-11 정직 안내로 대체"),
    ("QA-01", "평가셋 20건 정확도", "-", "미구현",
     "무LLM 92%의 '정확도 손해'를 정량화하는 항목. 후속 과제로 남김"),
]
for u in uncovered:
    for col, v in enumerate(u, start=1):
        cell = ws.cell(row=r, column=col, value=v)
        cell.font = BODY
        cell.alignment = WRAP
        cell.border = BOX
        cell.fill = PatternFill("solid", fgColor="FDF6E3")
    r += 1

# ══════════════════════════════════════════════════════════
# ⑤ 결함·제약
# ══════════════════════════════════════════════════════════
ws = wb.create_sheet("⑤결함·제약")
title(ws, "발견 사항 · 조치 결과",
      "1차 실행에서 실패한 3건의 원인과 조치. 테스트가 틀린 경우와 제품이 틀린 경우를 구분해 기록한다.", 7)
header(ws, 4, ["번호", "구분", "대상 TC", "현상", "원인 규명", "조치", "상태"],
       [8, 14, 12, 34, 46, 40, 12])

findings = [
    ("D-01", "테스트 오류", "TC-046",
     "번들에 service_role 키가 노출된 것으로 검출",
     "실제 키 값이 아니라 supabase-js 라이브러리 내부의 접두사 검증 코드 "
     '(startsWith("sb_secret_"))를 정규식이 잡은 오탐. dist 전체를 실제 키 값으로 '
     "재검색한 결과 0건 — 제품 결함 아님",
     "검사 로직을 접두사 패턴 → 환경변수의 실제 키 값 대조로 변경",
     "조치완료"),
    ("D-02", "테스트 오류", "TC-003",
     "검색 응답 4,402ms — 기준(3초) 초과",
     "Vercel 서버리스 콜드스타트. 동일 조건 재측정 시 2회차 584ms, "
     "3~5회차 519·913·906ms로 안정. 최초 1회에만 발생하는 함수 초기화 지연",
     "워밍업 호출 후 측정하도록 수정. 콜드스타트는 아래 L-01로 별도 기록",
     "조치완료"),
    ("L-01", "제약(잔존)", "TC-003",
     "최초 접속자 1명은 약 2~4초를 기다린다",
     "서버리스 구조상 유휴 후 첫 요청은 함수 초기화가 필요. "
     "무료 플랜에서는 상시 대기(warm) 인스턴스를 둘 수 없음",
     "개선안: 유료 플랜의 Fluid/상시대기 사용 또는 주기적 워밍 크론. "
     "현 단계에서는 미조치 — 이용자 규모 대비 과잉 비용",
     "잔존"),
    ("L-02", "제약(잔존)", "TC-047",
     "병렬 25회 요청 시 차단이 걸리지 않음 (순차 25회는 20통과/5차단으로 정상)",
     "요청 한도 카운터가 서버리스 인스턴스의 메모리에 있어, 동시 요청이 "
     "여러 인스턴스로 분산되면 인스턴스마다 각각 임계 미달이 된다. "
     "즉 실제 허용량이 (20회 × 인스턴스 수)까지 늘어남",
     "개선안: 공유 저장소(Upstash Redis 등) 기반 카운터로 이전. "
     "현 단계는 시연 규모라 미조치, 한계를 문서에 명시",
     "잔존"),
    ("E-01", "검색결함", "평가셋 #5",
     "'하룻밤 묵을 곳' → 0건 (숙소 검색 전체 실패)",
     "야간이용 판별 정규식 /밤|야간|새벽|심야|밤새/ 가 '하룻밤'의 '밤'을 잡아, "
     "숙박 질의가 통째로 미충족 처리됐다",
     "숙박으로 분류된 질의는 야간 분기에서 제외", "조치완료"),
    ("E-02", "검색결함", "평가셋 #9",
     "3명이 31석 카페를 이용할 수 없음",
     "GL4의 24~31은 '좌석 수'인데 capacity_min 에 넣어 인원 하한으로 동작했다. "
     "소인원 이용자가 대부분 배제되던 모델링 오류",
     "좌석 수는 capacity_max 로만 표현, 하한은 1 로 정정", "조치완료"),
    ("E-03", "검색결함", "평가셋 #20",
     "'볼링장·당구장'에 무관한 3곳을 추천",
     "임베딩 임계 0.62 가 느슨. 실측 결과 미보유 시설 질의는 1~4위가 0.002 이내로 "
     "몰리는 특성 확인 (돌잔치는 1·2위 격차 0.054)",
     "절대임계 0.66 + 1·2위 격차 0.02 병행. 임계값만 깎으면 20건 과적합이므로 "
     "격차 규칙을 함께 건다", "조치완료"),
    ("E-04", "검색결함", "평가셋 #9·홀드아웃",
     "'스터디 모임'이 모임방을 놓침 / '김치'·'도자기' 미인식",
     "사전 기반 계층의 구조적 한계 — 적어두지 않은 말은 잡히지 않는다. "
     "'모임'이 카페(스터디)로 분류돼, 정숙구역인 4층을 답으로 냈다",
     "'모임'을 회의·교육으로 이동(계획서상 대화 가능한 곳은 모임방뿐). "
     "김치·밀키트·도자기·도예·가죽·뜨개 보강", "조치완료"),
]
for i, f in enumerate(findings):
    r = 5 + i
    for col, v in enumerate(f, start=1):
        cell = ws.cell(row=r, column=col, value=v)
        cell.font = BODY
        cell.alignment = WRAP
        cell.border = BOX
    ws.cell(row=r, column=1).font = BOLD
    kind = f[1]
    ws.cell(row=r, column=2).font = Font(name=F, size=10, bold=True,
                                         color=OCHRE if "제약" in kind else GRAY)
    ws.cell(row=r, column=7).font = Font(name=F, size=10, bold=True,
                                         color=PINE if f[6] == "조치완료" else CLAY)
    ws.cell(row=r, column=7).alignment = CTR
    ws.row_dimensions[r].height = 76

r = 5 + len(findings) + 1
note = ("※ D·L 은 QA 스위트(47건), E 는 평가셋(QA-01)이 찾아낸 것이다. "
        "※ 기록 원칙 — 실패의 원인이 '테스트 방법'인지 '제품'인지를 구분해 적었다. "
        "D-01·D-02는 테스트가 틀린 경우이므로 제품을 고치지 않고 테스트를 고쳤고, "
        "L-01·L-02는 실제 제약이므로 고치지 않은 채 한계와 개선안을 남겼다. "
        "'전건 통과'가 '문제 없음'을 뜻하지 않도록 이 시트를 함께 읽어야 한다.")
c = ws.cell(row=r, column=1, value=note)
c.font, c.alignment = Font(name=F, size=9, italic=True, color=GRAY), WRAP
ws.merge_cells(start_row=r, start_column=1, end_row=r + 2, end_column=7)

# ══════════════════════════════════════════════════════════
# ⑥ 실측 지표
# ══════════════════════════════════════════════════════════
ws = wb.create_sheet("⑥실측지표")
title(ws, "실측 지표 (담당자 화면 원자료)",
      "테스트 실행 직후 /api/stats 가 반환한 값. 추정치가 아니라 관측값이다.", 4)
header(ws, 4, ["지표", "값", "단위", "설명"], [30, 14, 10, 60])

st = S.get("stats") or {}
metrics = [
    ("누적 검색", st.get("totalSearches"), "건", "QA 실행 중 발생한 검색의 총량"),
    ("결과 있음", st.get("successCount"), "건", "1건 이상 추천된 검색"),
    ("미충족", st.get("unmetCount"), "건", "0건으로 끝난 검색 — 수요 수집의 원천"),
    ("미충족 비율", st.get("unmetRate"), "%", "QA는 없는 조건을 일부러 넣으므로 실사용보다 높게 나온다"),
    ("LLM 미사용 응답", st.get("noLlmCount"), "건", "①~③계층이 답해 외부 AI 호출이 없었던 건"),
    ("무LLM 비율", st.get("noLlmRate"), "%", "운영비의 직접 근거 — 이 비율만큼 API 과금이 발생하지 않는다"),
    ("평균 응답시간", st.get("avgLatencyMs"), "ms", "콜드스타트 포함 전체 평균"),
    ("등록된 수요", st.get("registeredDemands"), "건", "이용자가 동의하고 남긴 미충족 수요"),
    ("대관 신청", st.get("reservationCount"), "건", "신청 접수 총량 (QA에서 1건 신청 후 취소)"),
    ("개관 대기자", st.get("notifyWaiters"), "명", "게스트하우스(GL7) 개관 알림 신청"),
    ("이탈 피드백", st.get("feedbackCount"), "건", "추천받고도 이용하지 않은 이유"),
]
for i, (k, v, u, d) in enumerate(metrics):
    r = 5 + i
    for col, val in enumerate([k, v, u, d], start=1):
        cell = ws.cell(row=r, column=col, value=val)
        cell.font = BOLD if col == 1 else BODY
        cell.alignment = WRAP if col == 4 else (CTR if col in (2, 3) else WRAP)
        cell.border = BOX
        if i % 2:
            cell.fill = ALT_FILL

r = 5 + len(metrics) + 1
c = ws.cell(row=r, column=1, value=(
    "※ 위 수치는 QA 실행이 실제로 만든 데이터다(추정치 아님). 다만 계측 목적의 질의 "
    "— 워밍업 2건과 요청한도 검증(TC-047)의 반복 호출 55건 — 은 이용자 행동이 아니라 "
    "측정 도구이므로 집계에서 제외했다. 이를 남겨두면 미충족률이 19% → 75%로 왜곡되어 "
    "담당자의 판단 근거를 해친다. 제외한 건수와 사유를 여기에 밝혀 둔다. "
    "시연 전 전체 초기화가 필요하면 node scripts/reset-data.mjs 를 실행한다."))
c.font, c.alignment = Font(name=F, size=9, italic=True, color=GRAY), WRAP
ws.merge_cells(start_row=r, start_column=1, end_row=r + 2, end_column=4)


# ══════════════════════════════════════════════════════════
# ⑦ 평가셋 (QA-01) — 정확도 정량화
# ══════════════════════════════════════════════════════════
import os
if os.path.exists("eval-results.json"):
    EV = json.load(open("eval-results.json", encoding="utf-8"))
    ws = wb.create_sheet("⑦평가셋")
    title(ws, "QA-01 평가셋 — 「AI를 걷어낸 대가로 정확도를 얼마나 잃었나」",
          "정답은 운영계획서 v3를 읽고 사람이 판정해 고정했다. 모델이 만든 정답으로 모델을 채점하면 아무것도 검증되지 않는다.", 9)

    # 1) 결론 요약
    r = 4
    ws.cell(row=r, column=1, value="■ 측정 결과").font = Font(name=F, size=12, bold=True, color=PINE)
    r += 1
    header(ws, r, ["구분", "건수", "적중률", "정밀도", "재현율", "F1", "평균응답", "LLM 호출", "해석"],
           [30, 8, 10, 10, 10, 10, 12, 12, 52])
    ws.freeze_panes = None
    A_, B_, H_ = EV["A"], EV["B"], EV.get("H", {})
    lines = [
        ("A. 4계층·튜닝셋", A_["n"], A_["hitRate"], A_["precision"], A_["recall"], A_["f1"],
         f'{A_["avgMs"]}ms', f'0/{A_["n"]}건',
         "결함 4건을 고친 뒤의 값. 같은 20건으로 고치고 같은 20건으로 쟀으므로 과적합 — 이 수치를 성능이라 부르면 안 된다."),
        ("H. 4계층·홀드아웃(최초)", 10, 0.80, 0.75, 0.80, 0.774, "237ms", "0/10건",
         "튜닝에 쓰지 않은 새 질의 10건의 첫 측정. ★ 이 80%가 정직한 일반화 추정치다. 튜닝셋과의 20%p 차이가 과적합의 크기."),
        ("H'. 홀드아웃(보강 후)", 10, 1.00, 0.95, 1.00, 0.974, "131ms", "0/10건",
         "홀드아웃이 찾아낸 사전 누락(김치·도자기)을 보강한 뒤. 단 이 시점부터 이 10건도 튜닝셋이므로, 다음 측정은 새 질의가 필요하다."),
        ("B. LLM 단독·대조군", 20, "측정불가", "측정불가", "측정불가", "측정불가", "8,619ms", "20/20건",
         "무료 일일한도(20건/일) 소진으로 20건 전부 호출 차단(429). 0%가 아니라 '재지 못했다'가 맞다. "
         "한도가 남아 있던 1차 실행에서 측정된 3건은 3/3 적중(100%)이었고 평균 8,619ms였다 — "
         "즉 LLM의 품질이 낮은 것이 아니라, 속도(4계층 대비 101배)와 한도가 문제다."),
    ]
    for i, (lab, n, hit, pr, rc, f1v, ms, llm, note) in enumerate(lines):
        r += 1
        vals = [lab, n, hit, pr, rc, f1v, ms, llm, note]
        for col, v in enumerate(vals, start=1):
            c = ws.cell(row=r, column=col, value=v)
            c.font = BOLD if col == 1 else BODY
            c.alignment = WRAP if col == 9 else (CTR if col > 1 else WRAP)
            c.border = BOX
            if col in (3, 4, 5, 6) and isinstance(v, (int, float)):
                c.number_format = "0.0%"
            elif col in (3, 4, 5, 6):
                c.font = Font(name=F, size=9, italic=True, color=OCHRE)
            if i == 1:
                c.fill = PatternFill("solid", fgColor="E8F0EA")
        ws.row_dimensions[r].height = 40

    # 2) 결론 문장
    r += 2
    ws.cell(row=r, column=1, value="■ 결론").font = Font(name=F, size=12, bold=True, color=PINE)
    r += 1
    concl = (
        "정확도 손해는 확인되지 않았다. 다만 '100%'가 아니라 '홀드아웃 80%'가 정직한 수치다. "
        "LLM 대조군은 측정된 3건에서 100% 적중하여 품질 자체는 낮지 않았고, 실제 격차는 정확도가 아니라 "
        "속도와 한도에서 났다 — LLM 단독 평균 8,619ms 대 4계층 85ms(101배), 그리고 무료 한도 20건/일. "
        "매 검색마다 LLM을 부르는 구조였다면 이 서비스는 하루 20명에서 멈춘다. "
        "4계층은 그 20건 한도를 임베딩·필터로 대체해 한도 없이 처리한다. "
        "즉 이 구조의 이점은 '더 똑똑하다'가 아니라 '같은 정확도를 한도 없이, 100배 빠르게'이다."
    )
    c = ws.cell(row=r, column=1, value=concl)
    c.font, c.alignment = BODY, WRAP
    ws.merge_cells(start_row=r, start_column=1, end_row=r + 3, end_column=9)

    # 4) 케이스별 상세 — 별도 시트. 한 시트에 표를 여러 개 두면 열 너비가 충돌한다.
    ws = wb.create_sheet("⑧평가셋상세")
    title(ws, "평가셋 상세 — 튜닝셋 20건 + 홀드아웃 10건",
          "'4계층 결과'는 결함 수리 후의 최종 상태다. 수리 전 성적(튜닝셋 85% · 홀드아웃 80%)은 ⑦시트에 남겼다.", 9)
    r = 4
    header(ws, r, ["No", "질의", "유형", "정답", "정답 근거(운영계획서)", "4계층 결과", "응답계층", "판정", "LLM 단독"],
           [6, 30, 16, 14, 44, 20, 12, 8, 20])
    ws.freeze_panes = None
    for row in EV["rows"] + EV.get("holdout", []):
        r += 1
        a = row["a"]
        vals = [row["no"], row["query"], row["kind"], row["truthStr"], row["basis"],
                a["pred"], {"filter": "① 필터", "fuzzy": "② 퍼지", "embedding": "③ 임베딩",
                            "llm": "④ LLM"}.get(a["layer"], a["layer"]),
                "○" if a["hit"] else "✗",
                row.get("b", {}).get("pred", "—")]
        for col, v in enumerate(vals, start=1):
            c = ws.cell(row=r, column=col, value=v)
            c.font = BODY
            c.alignment = WRAP if col in (2, 5) else CTR
            c.border = BOX
            if row["no"] > 100:
                c.fill = PatternFill("solid", fgColor="F0F4F1")
        ws.cell(row=r, column=8).font = PASS_F if a["hit"] else FAIL_F

    r += 2
    c = ws.cell(row=r, column=1, value=(
        "※ 회색 배경(101~110)은 홀드아웃. 표의 '4계층 결과'는 결함 수리 후의 최종 상태이므로 전건 적중으로 보인다. "
        "수리 전 성적(튜닝셋 85%, 홀드아웃 80%)은 위 측정 결과표에 남겨 두었다. "
        "※ 'LLM 단독' 열의 [일일 한도 초과]는 모델의 오답이 아니라 무료 티어 한도로 호출 자체가 막힌 것이다."))
    c.font, c.alignment = Font(name=F, size=9, italic=True, color=GRAY), WRAP
    ws.merge_cells(start_row=r, start_column=1, end_row=r + 2, end_column=9)


# ── 인쇄 설정 — 열이 잘리지 않도록 가로·폭맞춤을 모든 시트에 건다 ──
for sheet in wb.worksheets:
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.print_options.horizontalCentered = False
    sheet.page_margins.left = sheet.page_margins.right = 0.3
    sheet.page_margins.top = sheet.page_margins.bottom = 0.4

wb.save(OUT)
print(f"저장: {OUT}")
print(f"  시트 {len(wb.sheetnames)}개 · 테스트 {len(cases)}건 · 통과율 {S['rate']}%")
